#!/usr/bin/env python3
"""
apply_tool_modes.py — Tools Per-Division Tracking Mode, Phase 6 (set-mode intake).

Reads the operator-filled "Tools - Set Tracking Mode.xlsx" and applies each
(item, division) tracking mode to `inventory_item_divisions.tool_tracking_mode`,
then auto-applies Trading = bulk for every tool item held in the Trading
division. This is the bulk-apply tool that sets Trading=bulk / Maintenance=
serialized (etc.) at scale.

The write is an UPSERT on (item_id, division_id): 138 tool (item,division)
pairs hold stock with no assignment row yet, so the intake must be able to
INSERT a fresh override row, not only UPDATE. An override is written ONLY when
the desired mode differs from the item's category default — a mode that matches
the category is left inheriting (NULL), keeping the catalog clean. Every write
is pre-checked against on-hand stock so an effective-mode flip on a stocked
(item,division) is reported and skipped rather than hitting the DB guard
(migration 20260831001800), which is the transactional backstop. Exception
(migration 20260831002200): serialized -> bulk with ONLY bulk qty and no serial
units is the corrective, allowed flip (the counted-qty inventory load left tools
holding bulk qty under the serialized category default) — it proceeds instead of
being skipped.

Safety:
  * default is DRY-RUN — computes and reports, writes nothing.
  * --apply commits; --apply --rollback exercises the full write path
    (guards included) then rolls back — for verification without persisting.
  * --db staging (default) or --db newprod (reads NEW_DB_URL from
    supabase/.temp/migrate.env). The per-division feature must be deployed to
    the target first (column + tool_effective_mode + guards).

Usage:
  python scripts/apply_tool_modes.py                         # dry-run, staging
  python scripts/apply_tool_modes.py --apply                 # commit, staging
  python scripts/apply_tool_modes.py --db newprod --apply    # commit, new-prod
  python scripts/apply_tool_modes.py --xlsx "path/to.xlsx"   # custom file
"""
import argparse
import os
import re
import sys
from collections import defaultdict

import psycopg2
import openpyxl

DEFAULT_XLSX = r"D:\MMS\docs\Inventory Counts\Tools - Set Tracking Mode.xlsx"
MIGRATE_ENV = r"D:\MMS\supabase\.temp\migrate.env"

# Staging pooler host/user are NOT secrets (the project ref is in the committed
# supabase/config.toml). The password comes from the gitignored migrate.env
# (STAGING_DB_PASSWORD) or the environment — never hardcoded in this file.
STAGING_BASE = dict(
    host="aws-1-ap-south-1.pooler.supabase.com", port=5432,
    user="postgres.mwvblpgbgxipvrevkeff", dbname="postgres",
)

# Accepted values in the "Tracking Mode" column (case-insensitive).
MODE_ALIASES = {"serialized": "serialized", "serial": "serialized", "bulk": "bulk", "qty": "bulk"}


def _read_env(path=MIGRATE_ENV):
    env = {}
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            m = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.+?)\s*$", line)
            if m:
                env[m.group(1)] = m.group(2).strip().strip("'").strip('"')
    return env


def connect(target: str):
    env = _read_env()
    if target == "newprod":
        url = env.get("NEW_DB_URL")
        if not url:
            sys.exit("NEW_DB_URL not found in " + MIGRATE_ENV)
        return psycopg2.connect(url, connect_timeout=30)
    pw = env.get("STAGING_DB_PASSWORD") or os.environ.get("STAGING_DB_PASSWORD")
    if not pw:
        sys.exit("STAGING_DB_PASSWORD not found in " + MIGRATE_ENV + " (or environment)")
    return psycopg2.connect(**STAGING_BASE, password=pw, connect_timeout=30)


def norm(s):
    return str(s or "").strip().lower()


def load_catalog(cur):
    """Return lookup structures resolving the Excel's text columns to ids."""
    # Category path (same ' > ' join the generator used) for tool categories.
    cur.execute("""
        WITH RECURSIVE tree AS (
          SELECT id, name_en, parent_id, type, tool_tracking_mode, name_en::text AS path
            FROM inventory_categories WHERE parent_id IS NULL
          UNION ALL
          SELECT ic.id, ic.name_en, ic.parent_id, ic.type, ic.tool_tracking_mode,
                 t.path || ' > ' || ic.name_en
            FROM inventory_categories ic JOIN tree t ON ic.parent_id = t.id)
        SELECT id, path, tool_tracking_mode FROM tree WHERE type = 'tools'
    """)
    catpath, catmode = {}, {}
    for cid, path, mode in cur.fetchall():
        catpath[cid] = path
        catmode[cid] = mode

    # Tool items keyed by (category-path, name) — the Excel's natural key.
    cur.execute("""
        SELECT it.id, it.category_id, it.name_en
        FROM inventory_items it
        JOIN inventory_categories ic ON ic.id = it.category_id AND ic.type = 'tools'
    """)
    by_key = defaultdict(list)      # (catpath_lower, name_lower) -> [item_id]
    item_cat = {}                   # item_id -> category_id
    for iid, cid, name in cur.fetchall():
        by_key[(norm(catpath.get(cid, "")), norm(name))].append(iid)
        item_cat[iid] = cid

    # Divisions by name.
    cur.execute("SELECT id, name FROM company_divisions")
    div_by_name = {norm(n): i for i, n in cur.fetchall()}

    # Existing assignment/override rows.
    cur.execute("SELECT item_id, division_id, tool_tracking_mode FROM inventory_item_divisions")
    overrides = {}
    for iid, did, mode in cur.fetchall():
        overrides[(iid, did)] = mode
    existing = set(overrides.keys())

    # On-hand stock per (item, division): serial units and bulk qty.
    units = defaultdict(int)
    cur.execute("""
        SELECT item_id, division_id, count(*) FROM tool_asset_units
        WHERE division_id IS NOT NULL AND status <> 'retired'
        GROUP BY item_id, division_id
    """)
    for iid, did, n in cur.fetchall():
        units[(iid, did)] = int(n)
    qty = defaultdict(int)
    cur.execute("""
        SELECT v.item_id, sc.division_id, COALESCE(sum(f.remaining_qty), 0)
        FROM inventory_item_brand_variants v
        JOIN fifo_cost_layers f ON f.brand_variant_id = v.id AND f.remaining_qty > 0
        JOIN warehouse_sub_containers sc ON sc.id = f.sub_container_id
        JOIN inventory_items it ON it.id = v.item_id
        JOIN inventory_categories ic ON ic.id = it.category_id AND ic.type = 'tools'
        WHERE sc.division_id IS NOT NULL
        GROUP BY v.item_id, sc.division_id
    """)
    for iid, did, q in cur.fetchall():
        qty[(iid, did)] = int(q or 0)

    return dict(catmode=catmode, by_key=by_key, item_cat=item_cat,
                div_by_name=div_by_name, overrides=overrides, existing=existing,
                units=units, qty=qty)


def plan_write(cat, item_id, division_id, desired):
    """Decide the action for one (item,division,desired-mode).

    Returns (op, target_override, effective_change, stock) where op is one of
    'insert' | 'update' | 'nochange' | 'skip_stock'.
    """
    key = (item_id, division_id)
    cat_mode = cat["catmode"].get(cat["item_cat"][item_id])
    target = None if desired == cat_mode else desired      # None = inherit category
    row_exists = key in cat["existing"]
    current = cat["overrides"].get(key) if row_exists else "__ABSENT__"

    if row_exists:
        needs = (current != target)
    else:
        needs = (target is not None)
    if not needs:
        return ("nochange", target, False, (0, 0))

    eff_now = current if (row_exists and current is not None) else cat_mode
    eff_change = (desired != eff_now)
    if eff_change:
        u, q = cat["units"].get(key, 0), cat["qty"].get(key, 0)
        # Mirror the DB guard (migration 20260831002200): serialized -> bulk when
        # the division holds ONLY bulk qty and NO serial units is the corrective,
        # allowed flip (the qty is already bulk-shaped). Every other stocked flip
        # is skipped so it doesn't hit the guard.
        blocked = (u > 0 or q > 0) and not (desired == "bulk" and u == 0)
        if blocked:
            return ("skip_stock", target, True, (u, q))
    return (("update" if row_exists else "insert"), target, eff_change, (0, 0))


UPSERT = """
    INSERT INTO inventory_item_divisions (item_id, division_id, category_id, tool_tracking_mode)
    VALUES (%s, %s, %s, %s)
    ON CONFLICT (item_id, division_id) DO UPDATE SET tool_tracking_mode = EXCLUDED.tool_tracking_mode
"""


def main():
    ap = argparse.ArgumentParser(description="Apply per-division tool tracking modes from the intake Excel.")
    ap.add_argument("--db", choices=["staging", "newprod"], default="staging")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--apply", action="store_true", help="commit writes (default: dry-run)")
    ap.add_argument("--rollback", action="store_true", help="with --apply: execute then rollback (test)")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit("Excel not found: " + args.xlsx)

    conn = connect(args.db)
    conn.autocommit = False
    cur = conn.cursor()
    cat = load_catalog(cur)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}

    def col(r, name):
        i = hdr.get(name)
        return r[i] if i is not None and i < len(r) else None

    report = defaultdict(list)
    actions = []           # (op, item_id, division_id, target, label)
    seen_pairs = set()     # (item_id, division_id) covered by an Excel row

    for r in rows[1:]:
        if not r or not any(x is not None for x in r):
            continue
        catp = str(col(r, "Category Path") or "").strip()
        name = str(col(r, "Item Name") or "").strip()
        division = str(col(r, "Division") or "").strip()
        in_sys = str(col(r, "In System") or "").strip()
        mode_raw = norm(col(r, "Tracking Mode (serialized/bulk)"))
        label = f"{catp} / {name} [{division}]"

        if not mode_raw:
            report["blank"].append(label)
            continue
        desired = MODE_ALIASES.get(mode_raw)
        if desired is None:
            report["invalid_mode"].append(f"{label} = '{mode_raw}'")
            continue

        item_ids = cat["by_key"].get((norm(catp), norm(name)))
        if not item_ids:
            report["unresolved_item_expected" if in_sys.lower().startswith("no")
                   else "unresolved_item"].append(label)
            continue
        if len(item_ids) > 1:
            report["ambiguous_item"].append(f"{label} -> {len(item_ids)} matches")
            continue
        item_id = item_ids[0]
        division_id = cat["div_by_name"].get(norm(division))
        if not division_id:
            report["unresolved_division"].append(label)
            continue

        seen_pairs.add((item_id, division_id))
        op, target, _eff, stock = plan_write(cat, item_id, division_id, desired)
        if op == "nochange":
            report["nochange"].append(label)
        elif op == "skip_stock":
            report["skip_stock"].append(f"{label}: {stock[0]} unit(s), {stock[1]} qty on hand")
        else:
            actions.append((op, item_id, division_id, target, label))
            report[f"plan_{op}"].append(f"{label} -> {desired}"
                                        + ("" if target else " (inherits category)"))

    # ── Trading auto-bulk: every tool item held in Trading -> effective bulk ──
    trading_id = cat["div_by_name"].get("trading")
    if trading_id:
        cur.execute("""
            SELECT DISTINCT it.id FROM inventory_items it
            JOIN inventory_categories ic ON ic.id = it.category_id AND ic.type = 'tools'
            LEFT JOIN inventory_item_divisions d ON d.item_id = it.id AND d.division_id = %s
            LEFT JOIN inventory_item_brand_variants v ON v.item_id = it.id
            LEFT JOIN fifo_cost_layers f ON f.brand_variant_id = v.id AND f.remaining_qty > 0
            LEFT JOIN warehouse_sub_containers sc ON sc.id = f.sub_container_id AND sc.division_id = %s
            WHERE d.item_id IS NOT NULL OR sc.id IS NOT NULL
        """, (trading_id, trading_id))
        for (item_id,) in cur.fetchall():
            if (item_id, trading_id) in seen_pairs:
                continue  # an explicit Excel row already covered it (shouldn't happen — Trading is dropped)
            op, target, _eff, stock = plan_write(cat, item_id, trading_id, "bulk")
            if op == "nochange":
                report["trading_nochange"].append(item_id)
            elif op == "skip_stock":
                report["trading_skip_stock"].append(f"{item_id}: {stock[0]} unit(s), {stock[1]} qty")
            else:
                actions.append((op, item_id, trading_id, target, f"[Trading auto-bulk] {item_id}"))
                report[f"trading_plan_{op}"].append(item_id)

    # ── Execute ───────────────────────────────────────────────────────────────
    applied, failed = 0, []
    if args.apply and actions:
        for op, item_id, division_id, target, label in actions:
            try:
                cur.execute(UPSERT, (item_id, division_id, cat["item_cat"][item_id], target))
                applied += 1
            except Exception as e:
                conn.rollback()
                failed.append(f"{label}: {str(e).splitlines()[0]}")
        if failed:
            # A guard fired on something the pre-check missed (e.g. concurrent
            # stock change). Abort the whole batch so nothing partial persists.
            conn.rollback()
            applied = 0
        elif args.rollback:
            conn.rollback()
        else:
            conn.commit()
    else:
        conn.rollback()
    conn.close()

    # ── Report ────────────────────────────────────────────────────────────────
    def section(title, key, show=8):
        items = report.get(key, [])
        if not items:
            return
        print(f"\n{title}: {len(items)}")
        for x in items[:show]:
            print("   -", x)
        if len(items) > show:
            print(f"   … +{len(items) - show} more")

    mode = "APPLY" + (" (ROLLED BACK)" if args.rollback else "") if args.apply else "DRY-RUN"
    print("=" * 68)
    print(f"apply_tool_modes — {mode} — target={args.db}")
    print(f"Excel: {args.xlsx}")
    print("=" * 68)
    print(f"Excel rows planned:  insert={len(report.get('plan_insert', []))}  "
          f"update={len(report.get('plan_update', []))}")
    print(f"Trading auto-bulk:   insert={len(report.get('trading_plan_insert', []))}  "
          f"update={len(report.get('trading_plan_update', []))}  "
          f"nochange={len(report.get('trading_nochange', []))}  "
          f"skip_stock={len(report.get('trading_skip_stock', []))}")
    if args.apply:
        print(f"WRITES {'rolled back' if args.rollback else 'committed'}: {applied}"
              + (f"   FAILED: {len(failed)}" if failed else ""))
    section("Excel — planned INSERT", "plan_insert")
    section("Excel — planned UPDATE", "plan_update")
    section("Skipped (holds stock — empty first)", "skip_stock")
    section("No change (already correct)", "nochange", 4)
    section("Unfilled 'Tracking Mode' (skipped)", "blank", 4)
    section("Invalid mode value", "invalid_mode")
    section("Unresolved item — NOT in system (expected)", "unresolved_item_expected", 4)
    section("Unresolved item — IN system (INVESTIGATE)", "unresolved_item")
    section("Ambiguous item (multiple matches)", "ambiguous_item")
    section("Unresolved division", "unresolved_division")
    section("Trading — skipped (holds stock)", "trading_skip_stock")
    if failed:
        print(f"\nFAILED writes (batch rolled back): {len(failed)}")
        for x in failed[:8]:
            print("   -", x)
    print()


if __name__ == "__main__":
    main()
