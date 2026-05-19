# generate_inventory_migration.py
# --------------------------------
# Reads C:/Users/IT/Downloads/Inventory.xlsx and generates a Supabase SQL migration
# file at supabase/migrations/20260518120000_inventory_categories_and_items.sql
#
# Run from project root:
#     python scripts/migrate/generate_inventory_migration.py

import re
import uuid
import os
import math
import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
XLSX_PATH = r"C:\Users\IT\Downloads\Inventory.xlsx"
OUTPUT_SQL = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),  # scripts/migrate/
    "..", "..",                                   # project root
    "supabase", "migrations",
    "20260518120000_inventory_categories_and_items.sql",
)
OUTPUT_SQL = os.path.normpath(OUTPUT_SQL)

# ---------------------------------------------------------------------------
# Category definitions  (name_en, name_ar, type, sku_prefix)
# ---------------------------------------------------------------------------
CATEGORIES = [
    ("AC – Split – Piston",           "مكيف سبلت – بيستون",     "products",    "AC-SPL-PST"),
    ("AC – Split – Rotary",           "مكيف سبلت – روتاري",      "products",    "AC-SPL-ROT"),
    ("AC – Split – Inverter",         "مكيف سبلت – إنفرتر",      "products",    "AC-SPL-INV"),
    ("AC – Window – Piston",          "مكيف شباك – بيستون",     "products",    "AC-WIN-PST"),
    ("AC – Window – Rotary",          "مكيف شباك – روتاري",     "products",    "AC-WIN-ROT"),
    ("AC – Window – Inverter",        "مكيف شباك – إنفرتر",     "products",    "AC-WIN-INV"),
    ("AC – Stand – Rotary",           "مكيف ستاند – روتاري",     "products",    "AC-STD-ROT"),
    ("AC – Stand – Copeland",         "مكيف ستاند – كوبلاند", "products",    "AC-STD-COP"),
    ("AC – Stand – Inverter",         "مكيف ستاند – إنفرتر",     "products",    "AC-STD-INV"),
    ("AC – Floor Ceiling – Inverter", "مكيف أرضي سقفي – إنفرتر", "products", "AC-FC-INV"),
    ("Water Cooler – Piston",    "مبرد مياه – بيستون",    "products",    "WC-PST"),
    ("Water Cooler – Rotary",    "مبرد مياه – روتاري",    "products",    "WC-ROT"),
    ("Water Heater",                  "سخان مياه",                                                 "products",    "WH"),
    ("Water Pump – Circulation",                 "مضخة دوران",              "spare-parts", "WP-CIRC"),
    ("Water Pump – Potable – In-line Normal", "مضخة شرب – خط طبيعي", "spare-parts", "WP-POT-ILN"),
    ("Water Pump – Potable – In-line Silent", "مضخة شرب – صامتة",           "spare-parts", "WP-POT-ILS"),
    ("Water Pump – Potable – Submersible",    "مضخة شرب – غاطسة",           "spare-parts", "WP-POT-SUB"),
    ("Water Pump – Sewage",      "مضخة صرف صحي",                                              "spare-parts", "WP-SEW"),
    ("Water Pump – Rainwater",   "مضخة مياه أمطار",                            "spare-parts", "WP-RAIN"),
    ("Electrical – Sockets & Switches", "كهرباء – مقابس ومفاتيح", "spare-parts", "EL-SOCK"),
    ("Electrical – MCB & ELCB",         "كهرباء – قواطع",                    "spare-parts", "EL-MCB"),
    ("Electrical – LED Lighting",       "كهرباء – إضاءة LED",                "spare-parts", "EL-LED"),
    ("Electrical – Timer Switches",     "كهرباء – مؤقتات",              "spare-parts", "EL-TIMER"),
    ("Electrical – Isolators & Outlets","كهرباء – معزلات",              "spare-parts", "EL-ISO"),
    ("Plumbing – Valves",               "سباكة – صمامات",                    "spare-parts", "PL-VALV"),
    ("Plumbing – Hoses & Fittings",     "سباكة – خراطيم وتوصيلات", "spare-parts", "PL-HOSE"),
    ("Plumbing – Float Switches & Controls", "سباكة – عوامات وضغط", "spare-parts", "PL-CTRL"),
]

# ---------------------------------------------------------------------------
# AC regex  -- group1=ac_type  group2=brand  group3=ton  group4=compressor
# Note: in the Excel the separator is plain hyphen-minus (-)
# ---------------------------------------------------------------------------
AC_RE = re.compile(
    r'AC\s*[-–]\s*(Split|Window|Stand|Floor\s*Ceiling)\s*[-–]\s*(.+?)\s*[-–]\s*([\d.]+)\s*[Tt]on\s*[-–]\s*(Piston|Rotary|Inverter|Copeland)',
    re.IGNORECASE,
)

# en-dash character used throughout
ENDASH = "–"

# ---------------------------------------------------------------------------
# Hardcoded Water Cooler rows  (compressor_type, ton, brand, cost)
# ---------------------------------------------------------------------------
WC_ITEMS = [
    ("Rotary", "1.5", "Alfacool", 1299.40),
    ("Rotary", "2",   "Alfacool", 1500.15),
    ("Piston", "1.5", "Alfacool", 2153.50),
    ("Rotary", "1.5", "Hommer",   1250.00),
    ("Rotary", "2",   "Hommer",   1650.00),
    ("Piston", "2",   "Hommer",   2336.00),
    ("Rotary", "3",   "Hommer",   2700.00),
]

# ---------------------------------------------------------------------------
# Hardcoded Water Heater rows  (brand, size_en, size_ar, cost)
# ---------------------------------------------------------------------------
WH_ITEMS = [
    ("Alfaheat",              "80 Gallon", "سخان 80 جالون", 1405.41),
    ("Bradford White",        "80 Gallon", "سخان 80 جالون", 2250.00),
    ("American Water Heater", "80 Gallon", "سخان 80 جالون", 2200.00),
]

# ---------------------------------------------------------------------------
# Hardcoded Water Pump rows
# (cat_name_en, item_name_en, item_name_ar, item_sku, brand, variant_code, cost)
# ---------------------------------------------------------------------------
WP_ITEMS = [
    ("Water Pump – Circulation", "Circulation Pump", "مضخة دوران", "WP-CIRC-PUMP", "DAB",     "VA-65/180",       410.00),
    ("Water Pump – Circulation", "Circulation Pump", "مضخة دوران", "WP-CIRC-PUMP", "SPCO",    "FGD 25-60-180 A", 310.00),
    ("Water Pump – Circulation", "Circulation Pump", "مضخة دوران", "WP-CIRC-PUMP", "SPCO",    "SCP 32-8S",       320.00),
    ("Water Pump – Potable – In-line Silent", "1 HP",   "1 حصان",   "WP-ILS-1HP",  "DAB", "KI-30/90",  750.00),
    ("Water Pump – Potable – In-line Silent", "1.3 HP", "1.3 حصان", "WP-ILS-13HP", "DAB", "KI-30/120", 1050.00),
    ("Water Pump – Potable – Submersible",    "SCM 4",      "SCM 4",      "WP-SUB-SCM4", "SPCO",    "SCM 4",      780.00),
    ("Water Pump – Potable – Submersible",    "SP 22 Dirt", "SP 22 Dirt", "WP-SUB-SP22", "Karcher", "SP 22 Dirt", 823.81),
    ("Water Pump – Potable – Submersible",    "SP7",        "SP7",        "WP-SUB-SP7",  "Karcher", "SP7",        0.00),
]

# ---------------------------------------------------------------------------
# Electrical / Plumbing keyword classifiers
# ---------------------------------------------------------------------------
ELEC_CLASSIFY = [
    (["13AMP Socket", "Switch One Gang", "Switch Two Gang", "Switch Three Gang", "3-Pin Top", "Fuse 13AMP"],
     "Electrical – Sockets & Switches"),
    (["MCB"],                                "Electrical – MCB & ELCB"),
    (["ELCB"],                               "Electrical – MCB & ELCB"),
    (["LED Spot Light", "LED Lamp", "LED COB"], "Electrical – LED Lighting"),
    (["Timer Switch"],                       "Electrical – Timer Switches"),
    (["Isolator", "Outlet", "Switch Double Pole", "13AMP Spare Unit"],
     "Electrical – Isolators & Outlets"),
]

PLUMB_CLASSIFY = [
    (["Non-Return Valve", "Float Valve", "Angle Valve", "Water valve"], "Plumbing – Valves"),
    (["Magic Hose", "Pop-Up Waste", "Shattaf", "Drainage Cleaner"],    "Plumbing – Hoses & Fittings"),
    (["Float Switch", "Pressure Control"],                              "Plumbing – Float Switches & Controls"),
]

# ---------------------------------------------------------------------------
# Skip patterns -- rows whose Name contains any of these are ignored
# ---------------------------------------------------------------------------
SKIP_FRAGMENTS = [
    # AC rows are handled by the AC regex parser -- do NOT skip them here
    # Water Cooler / Heater are handled by hardcoded mappings
    "Water Cooler", "Water Heater",
    # Pumps already handled by hardcoded WP_ITEMS
    "Slient Pump", "Silent Pump",
    "Water Circulator",
    "Submersible", "SPCO", "SYPCO", "Karcher",
    # Water filter / consumable products not in scope
    "Water Filter", "CCK", "Alkawther", "Alfapure",
    # AC spare parts / consumables handled via service links
    "Air Deflector", "Capacitor", "Contactor", "Coil temp",
    "Water Heart", "Stainless Steel", "Shower",
    "Membrane", "CTO 2", "UDF 10", "PP 10", "PP 20",
    "Purewater", "Aqua", "Bradford", "American",
    "Alfaheat", "Hommer",
    "Electric Wire", "Relay",
]

# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------

def sq(val):
    # type: (str) -> str
    # Escape a string for SQL single-quoted literal.
    return val.replace("'", "''")


def cost_sql(cost):
    # Return SQL representation of a cost value (NULL if 0 or NaN).
    try:
        f = float(cost)
        if f == 0 or math.isnan(f):
            return "NULL"
        return str(f)
    except (TypeError, ValueError):
        return "NULL"


# ---------------------------------------------------------------------------
# Build category lookup:  name_en  ->  {id, name_ar, type, sku_prefix}
# ---------------------------------------------------------------------------
cat_map = {}
for _name_en, _name_ar, _cat_type, _sku_prefix in CATEGORIES:
    cat_map[_name_en] = {
        "id": str(uuid.uuid4()),
        "name_en": _name_en,
        "name_ar": _name_ar,
        "type": _cat_type,
        "sku": _sku_prefix,
    }

# ---------------------------------------------------------------------------
# Accumulators
# item_ids:  (cat_id, item_name_en)  ->  item_id
# ---------------------------------------------------------------------------
item_ids = {}
items_rows = []
variant_rows = []


def get_or_create_item(cat_name_en, item_name_en, item_name_ar, item_sku):
    # Return existing item_id or create new item row and return its id.
    cat = cat_map[cat_name_en]
    key = (cat["id"], item_name_en)
    if key in item_ids:
        return item_ids[key]
    iid = str(uuid.uuid4())
    item_ids[key] = iid
    items_rows.append({
        "id": iid,
        "category_id": cat["id"],
        "name_en": item_name_en,
        "name_ar": item_name_ar,
        "sku": item_sku,
    })
    return iid


def add_variant(item_id, brand, code, cost):
    variant_rows.append({
        "id": str(uuid.uuid4()),
        "item_id": item_id,
        "brand": brand,
        "code": code,
        "cost_price": cost_sql(cost),
    })


# ---------------------------------------------------------------------------
# 1. Water Cooler (hardcoded)
# ---------------------------------------------------------------------------
for _compressor, _ton, _brand, _cost in WC_ITEMS:
    _cat_name = "Water Cooler – " + _compressor
    _item_name_en = _ton + " Ton"
    _item_name_ar = _ton + " طن"   # Arabic: X طن
    _cat_sku = cat_map[_cat_name]["sku"]
    _item_sku = _cat_sku + "-" + _ton.replace(".", "_") + "T"
    _item_id = get_or_create_item(_cat_name, _item_name_en, _item_name_ar, _item_sku)
    add_variant(_item_id, _brand, "", _cost)

# ---------------------------------------------------------------------------
# 2. Water Heater (hardcoded)
# ---------------------------------------------------------------------------
_WH_CAT = "Water Heater"
for _brand, _size_en, _size_ar, _cost in WH_ITEMS:
    _item_sku = "WH-" + _size_en.replace(" ", "-")
    _item_id = get_or_create_item(_WH_CAT, _size_en, _size_ar, _item_sku)
    add_variant(_item_id, _brand, "", _cost)

# ---------------------------------------------------------------------------
# 3. Water Pumps (hardcoded)
# ---------------------------------------------------------------------------
for _cat_name_en, _item_name_en, _item_name_ar, _item_sku, _brand, _code, _cost in WP_ITEMS:
    _item_id = get_or_create_item(_cat_name_en, _item_name_en, _item_name_ar, _item_sku)
    add_variant(_item_id, _brand, _code, _cost)

# ---------------------------------------------------------------------------
# 4. Read Excel: process AC rows + Elec/Plumbing rows
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active
_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
_col_idx = {h: i + 1 for i, h in enumerate(_headers)}

_skipped = 0
_unclassified = []


def _process_ac_row(name, cost):
    # Returns True if the row was handled as an AC item.
    m = AC_RE.match(name)
    if not m:
        return False
    _ac_type_raw = m.group(1)
    _b = m.group(2).strip()
    _ton = m.group(3)
    _comp = m.group(4)

    # Normalise ac_type to title case with single space
    _ac_type = re.sub(r'\s+', ' ', _ac_type_raw).title()  # "Floor Ceiling", "Split", "Window", "Stand"

    # Capitalise compressor consistently
    _comp_lower = _comp.lower()
    if _comp_lower == "copeland":
        _comp_cap = "Copeland"
    elif _comp_lower == "inverter":
        _comp_cap = "Inverter"
    elif _comp_lower == "rotary":
        _comp_cap = "Rotary"
    else:
        _comp_cap = "Piston"

    _cat_name = "AC – " + _ac_type + " – " + _comp_cap

    if _cat_name not in cat_map:
        print("  [WARN] AC category not in map: " + repr(_cat_name) + "  from name=" + repr(name))
        return False

    _item_name_en = _ton + " Ton"
    _item_name_ar = _ton + " طن"   # طن
    _cat_sku = cat_map[_cat_name]["sku"]
    _item_sku = _cat_sku + "-" + _ton.replace(".", "_") + "T"

    _iid = get_or_create_item(_cat_name, _item_name_en, _item_name_ar, _item_sku)
    add_variant(_iid, _b, "", cost)
    return True


for _row in range(2, ws.max_row + 1):
    _is_group = ws.cell(_row, _col_idx["IsGroup"]).value
    if str(_is_group).lower() != "false":
        continue

    _name = ws.cell(_row, _col_idx["Name"]).value or ""
    _name = _name.strip()
    if not _name:
        continue

    _cost = ws.cell(_row, _col_idx["Cost"]).value

    # --- Skip list ---
    if any(_frag.lower() in _name.lower() for _frag in SKIP_FRAGMENTS):
        _skipped += 1
        continue

    # --- AC rows ---
    if _process_ac_row(_name, _cost):
        continue

    # --- Electrical ---
    _classified = False
    for _keywords, _cat_name in ELEC_CLASSIFY:
        if any(_kw.lower() in _name.lower() for _kw in _keywords):
            # Split on last en-dash (U+2013) to separate item description from brand
            if ENDASH in _name:
                _parts = _name.rsplit(ENDASH, 1)
                _item_name_en = _parts[0].strip()
                _brand = _parts[1].strip()
            else:
                _item_name_en = _name.strip()
                _brand = "Generic"
            _item_name_ar = _item_name_en  # no Arabic in source data
            _item_sku = (
                cat_map[_cat_name]["sku"]
                + "-"
                + re.sub(r"[^A-Z0-9]", "", _item_name_en.upper())[:12]
            )
            _iid = get_or_create_item(_cat_name, _item_name_en, _item_name_ar, _item_sku)
            add_variant(_iid, _brand, "", _cost)
            _classified = True
            break

    if _classified:
        continue

    # --- Plumbing ---
    for _keywords, _cat_name in PLUMB_CLASSIFY:
        if any(_kw.lower() in _name.lower() for _kw in _keywords):
            if ENDASH in _name:
                _parts = _name.rsplit(ENDASH, 1)
                _item_name_en = _parts[0].strip()
                _brand = _parts[1].strip()
            else:
                _item_name_en = _name.strip()
                _brand = "Generic"
            _item_name_ar = _item_name_en
            _item_sku = (
                cat_map[_cat_name]["sku"]
                + "-"
                + re.sub(r"[^A-Z0-9]", "", _item_name_en.upper())[:12]
            )
            _iid = get_or_create_item(_cat_name, _item_name_en, _item_name_ar, _item_sku)
            add_variant(_iid, _brand, "", _cost)
            _classified = True
            break

    if _classified:
        continue

    _unclassified.append(_name)

# ---------------------------------------------------------------------------
# 5. Generate SQL
# ---------------------------------------------------------------------------
_lines = []
_lines.append("-- Auto-generated by scripts/migrate/generate_inventory_migration.py")
_lines.append("-- DO NOT EDIT MANUALLY\n")
_lines.append("BEGIN;\n")

# --- inventory_categories ---
_lines.append("-- ============================================================")
_lines.append("-- inventory_categories  (27 rows)")
_lines.append("-- ============================================================")
_lines.append("INSERT INTO inventory_categories (id, name_en, name_ar, sku, type) VALUES")
_cat_values = []
for _name_en, _name_ar, _cat_type, _sku_prefix in CATEGORIES:
    _c = cat_map[_name_en]
    _cat_values.append(
        "  ('" + _c["id"] + "', '" + sq(_c["name_en"]) + "', '" + sq(_c["name_ar"])
        + "', '" + sq(_c["sku"]) + "', '" + _c["type"] + "')"
    )
_lines.append(",\n".join(_cat_values) + ";\n")

# --- inventory_items ---
_lines.append("-- ============================================================")
_lines.append("-- inventory_items  (" + str(len(items_rows)) + " rows)")
_lines.append("-- ============================================================")
_lines.append("INSERT INTO inventory_items (id, category_id, name_en, name_ar, sku) VALUES")
_item_values = []
for _it in items_rows:
    _item_values.append(
        "  ('" + _it["id"] + "', '" + _it["category_id"] + "', '"
        + sq(_it["name_en"]) + "', '" + sq(_it["name_ar"]) + "', '" + sq(_it["sku"]) + "')"
    )
_lines.append(",\n".join(_item_values) + ";\n")

# --- inventory_brand_variants ---
_lines.append("-- ============================================================")
_lines.append("-- inventory_brand_variants  (" + str(len(variant_rows)) + " rows)")
_lines.append("-- ============================================================")
_lines.append(
    "INSERT INTO inventory_brand_variants "
    "(id, item_id, brand, code, cost_price, selling_price, stock_level, incoming, average_cost, reserved_qty) VALUES"
)
_var_values = []
for _v in variant_rows:
    _var_values.append(
        "  ('" + _v["id"] + "', '" + _v["item_id"] + "', '" + sq(_v["brand"])
        + "', '" + sq(_v["code"]) + "', "
        + _v["cost_price"] + ", NULL, 0, 0, " + _v["cost_price"] + ", 0)"
    )
_lines.append(",\n".join(_var_values) + ";\n")

_lines.append("COMMIT;")

# ---------------------------------------------------------------------------
# 6. Write file
# ---------------------------------------------------------------------------
os.makedirs(os.path.dirname(OUTPUT_SQL), exist_ok=True)
with open(OUTPUT_SQL, "w", encoding="utf-8") as _f:
    _f.write("\n".join(_lines))

# ---------------------------------------------------------------------------
# 7. Summary
# ---------------------------------------------------------------------------
print("")
print("=" * 60)
print("Generated: " + OUTPUT_SQL)
print("  inventory_categories     : " + str(len(CATEGORIES)))
print("  inventory_items          : " + str(len(items_rows)))
print("  inventory_brand_variants : " + str(len(variant_rows)))
print("  Skipped (skip list)      : " + str(_skipped))
print("  Unclassified rows        : " + str(len(_unclassified)))
if _unclassified:
    print("\nUnclassified names:")
    for _n in _unclassified:
        print("  " + repr(_n))
print("=" * 60)

# Verify minimums
assert len(CATEGORIES) == 27, "Expected 27 categories, got " + str(len(CATEGORIES))
assert len(items_rows) >= 50, "Expected >=50 items, got " + str(len(items_rows))
assert len(variant_rows) >= 80, "Expected >=80 variants, got " + str(len(variant_rows))
assert ("Water Pump – Sewage" in cat_map), "Missing Water Pump - Sewage category"
assert ("Water Pump – Rainwater" in cat_map), "Missing Water Pump - Rainwater category"
print("\nAll assertions passed.")
