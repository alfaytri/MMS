# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

MATRIX = [
 ("Master Data", [
   ("Inventory", "/master-data/inventory", "inventory.catalog.view", [
     "List loads — categories, items, brands, origins render",
     "Create category + nested sub-category",
     "Create item with a brand + origin variant",
     "Edit item details (name, unit, category)",
     "Edit variant cost / selling price",
     "Archive item, variant, and category",
     "Reorder catalog tree (drag) persists after reload",
     "Category attributes — create / edit / archive definitions & options",
     "Set attribute values on an item",
     "Filter inventory by attribute chips",
     "Assign an item to one or more divisions",
     "Bulk import (.xlsx) — upload + validation errors surface",
     "Search finds items / brands / SKUs",
     "Tools & Assets category tracking mode (Bulk/Serialized) cascade + locked summary",
   ]),
   ("Warehouses", "/master-data/warehouses", "warehouse.access", [
     "Warehouses tab loads",
     "Create / edit a warehouse",
     "Assign Warehouse Responsible Person(s)",
     "Stock Overview — per-warehouse quantities (+ cost if permitted)",
     "Movements tab loads with correct history",
     "Projects — create / close, add disciplines & milestones",
     "Stock Value tab (permission-gated) shows valuation",
   ]),
   ("Users & Roles", "/master-data/users", "master_data.users.view / roles.view", [
     "Users list loads",
     "Create user — auth + profile + role assignment",
     "Edit user profile",
     "Assign / remove a role on a user",
     "Assign / remove a division on a user",
     "Roles tab — create a new role",
     "Edit role permissions in the permission tree (save persists)",
     "Delete a role",
     "* Delete a role used in a PO band -> it disappears from the band (cascade fix)",
     "* Rename a role -> the new name shows in the PO band (cascade fix)",
     "Non-admin role: displayed label is a name, never a raw UUID",
   ]),
   ("Audit Trail", "/master-data/audit-trail", "master_data.audit.view", [
     "Audit log loads",
     "Filter by module / user / date range",
     "Entry shows actor, action, before/after where relevant",
   ]),
   ("Admin - Approval Settings (PO Bands)", "/master-data/admin/approval-settings", "purchase.approvals.chain.manage", [
     "Company Default chain + tiers render",
     "Add Tier (min/max/roles) saves",
     "Edit a tier inline (amounts + roles) saves",
     "Delete a tier",
     "Create a division-specific chain + toggle active",
     "Archive a division chain -> falls back to Company Default",
     "* Empty a band (remove its only role) -> amber 'no approvers' warning shows",
     "* Submit a PO in an empty band -> blocked with a clear message (fail-closed)",
   ]),
   ("Admin - other", "/master-data/admin", "master_data.admin.view", [
     "Admin landing + sidebar links gate by permission",
     "Warehouses / Custody admin pages load & edit",
     "Repair vendors — create / edit / delete",
     "Credit-group approvals — approve / reject",
     "Credit groups — create / edit / delete",
     "Payment methods — rename cascades to existing records",
     "Workflow steps / groups — add / update / archive",
   ]),
   ("Suppliers", "/master-data/suppliers", "master_data.suppliers.view", [
     "Suppliers list loads",
     "Create a supplier (phone uses country-code input)",
     "Edit a supplier",
     "Supplier detail — bills / payments / aging tabs",
   ]),
   ("Customers", "/master-data/customers", "master_data.customers.view", [
     "Customers list loads",
     "Create a customer",
     "Edit a customer",
     "Toggle customer active / inactive",
     "Change customer type",
     "* Raise credit group directly -> blocked (must go through approval)",
     "Submit credit-group change -> approval flow",
     "Upload customer credit documents",
   ]),
 ]),
 ("Reports", [
   ("All financial reports", "/reports/*", "reports.* (per-report)", [
     "Financial Dashboard loads (finance section gated)",
     "Product Profitability loads + drill-down",
     "Product Cost loads",
     "Revenue & COGS loads (retro LC + consumption in COGS)",
     "Accounts Receivable loads",
     "Accounts Payable loads",
     "Cash & Cash Equivalents loads",
     "Profit & Loss loads (Stock-Value badge, currency labels)",
     "Consumption report loads (cost gated)",
     "Division filter re-scopes every report",
     "Date-range picker (calendar) filters",
     "Excel export (server-side /api/reports/excel) downloads",
   ]),
 ]),
 ("Purchase", [
   ("Purchase Orders", "/purchase/orders", "purchase.orders.view", [
     "PO list loads (respects 50-row cap / server search)",
     "Create a draft PO",
     "Create an RFQ (multi-supplier)",
     "Edit / amend a PO (line items, division)",
     "Save an RFQ quote",
     "Award an RFQ quote -> PO",
     "Submit PO for approval -> status pending_approval, correct tiers",
     "Cancel a PO",
     "Soft-delete a draft PO",
     "Record a supplier payment against the PO",
     "Change booked exchange rate",
     "Request a post-approval edit -> review -> mark used",
     "PO PDF opens (authorized) / 403 if not permitted",
     "Version snapshots (po-v1...) saved on submit/resubmit",
   ]),
   ("Approvals (Purchase)", "/purchase/approvals", "purchase.approvals.view", [
     "Pending approvals list loads",
     "Approve a step",
     "Reject a PO (with reason)",
     "Force-approve a step (Owner)",
     "Force-approve all steps",
     "Approved PO flips to approved; rejected returns to draft",
   ]),
   ("Receivals", "/purchase/receivals", "purchase.receivals.view", [
     "Receivals list loads",
     "Create & approve a receival (books FIFO layers + stock)",
     "Request a receival edit -> approve edit",
     "Self-serve receival edit",
     "Create a replacement receival",
     "Receival check + receival receipt PDFs open",
     "Retry / double-submit does NOT double-post (idempotency check)",
   ]),
   ("Bills", "/purchase/bills", "purchase.bills.view", [
     "Bills list loads",
     "Create a supplier bill (recompute is idempotent)",
     "Attach a payment to a bill",
     "Edit / delete a supplier payment",
     "Attach supplier invoice files",
     "Bill PDF opens",
   ]),
   ("Returns (Purchase)", "/purchase/returns", "purchase.returns.view", [
     "PO returns list loads",
     "Create a purchase return",
     "Dispatch a PO return; cancel dispatch",
     "Issue a debit note",
     "Resolve debit note as supplier credit",
     "Resolve debit note as replacement",
     "Direct status write blocked (guarded)",
   ]),
   ("Debit Notes", "/purchase/debit-notes", "purchase.debit_notes.view", [
     "Debit notes list loads",
     "Debit note PDF opens",
   ]),
   ("Aging Report (Purchase)", "/purchase/aging-report", "purchase.bills.view / aging.view", [
     "Payables aging loads",
     "Buckets + totals correct; filter works",
   ]),
   ("Shipments", "/purchase/shipments", "purchase.shipments.view", [
     "Shipments list loads",
     "Create a shipment",
     "Update shipment status",
     "Add a shipment event",
     "Archive a shipment",
   ]),
   ("Landed Costs", "/purchase/landed-costs", "purchase.landed_costs.view", [
     "Landed costs list loads",
     "Create a landed cost",
     "Apply / allocate to receival layers",
     "Revert an allocation",
     "Void a landed cost (COGS visibility in P&L)",
   ]),
   ("Dead Stock Report", "/purchase/dead-stock", "purchase.dead_stock.view", [
     "Dead stock report loads + filters",
   ]),
 ]),
 ("Sales", [
   ("Sale Orders", "/sales/orders", "sales.orders.view", [
     "SO list loads",
     "Create a quotation",
     "Create a confirmed SO (credit check at creation)",
     "Edit / amend an SO",
     "* Save as quotation -> Confirm -> re-runs credit gate (over-limit -> approval)",
     "Cancel an SO",
     "Resubmit an SO after rejection",
     "SO / quotation PDF opens",
   ]),
   ("Approvals (Sales)", "/sales/approvals", "sales.approvals.view", [
     "Sales approvals list loads (margin / credit)",
     "Approve / reject / force-approve",
     "Over-limit order routes here instead of auto-confirming",
   ]),
   ("SO Invoices", "/sales/invoices", "sales.invoices.view", [
     "Invoices list loads (tl_invoices, source=order)",
     "Generate an invoice from an SO",
     "Void an invoice",
     "Issue a credit note against an invoice",
     "Apply a credit note to an invoice",
     "Invoice PDF opens (authorized only)",
     "QuickBooks bulk sync (if used)",
   ]),
   ("Returns (Sales)", "/sales/returns", "sales.returns.view", [
     "Sales returns list loads",
     "Create a direct sales return",
     "Complete return inspection",
     "Restock a return",
     "Issue a credit note for the return",
     "Create a partial replacement (customer + inventory atomic)",
     "Record refund / store credit / disposition",
   ]),
   ("Warranties", "/sales/warranties", "sales.warranties.view", [
     "Warranty records list loads (origin snapshot shown)",
     "Warranty certificate PDF opens",
     "File a warranty claim",
     "Assess — cover / reject a claim",
     "Start resolution -> spawns warranty-flagged return",
     "Void a claim (releases coverage; cancels pending return)",
     "Partial claim qty + remaining-coverage tracking",
   ]),
   ("Deliveries", "/sales/deliveries", "sales.deliveries.view", [
     "Deliveries list loads",
     "Create a delivery from an SO",
     "Edit a draft delivery",
     "Complete delivery -> books stock + issues warranty records",
     "Auto-open warranty certificate on completion",
     "Cancel a delivery (restocks correctly)",
     "Delivery note PDF opens",
   ]),
   ("Credit Notes", "/sales/credit-notes", "sales.credit_notes.view", [
     "Credit notes list loads",
     "Credit note PDF opens",
     "Apply store credit (FIFO across CNs)",
   ]),
   ("Customer Statement", "/sales/customer-statement", "sales.invoices.view / customer_statement.view", [
     "Statement loads for a customer",
     "Statement PDF opens (authorized only — no cross-customer access)",
   ]),
   ("Aging Report (Sales)", "/sales/aging-report", "sales.invoices.view / aging.view", [
     "Receivables aging loads; buckets + filter correct",
   ]),
   ("Payments / Pending Payments", "/invoices/pending-payments", "payments.view", [
     "Pending payments list loads",
     "Record a customer payment",
     "Change booked exchange rate (SO)",
     "Attach / detach payment from invoice",
     "Create a payment plan (installments)",
     "Settle an installment",
   ]),
 ]),
 ("Operations", [
   ("Custody", "/warehouse/custody", "custody.view", [
     "Custody holdings list loads (per responsible person)",
     "Custody -> custody transfer (hand-out)",
     "Consume from custody",
     "Cost column hidden without custody.cost.view",
   ]),
   ("Consumption", "/consumption", "consumption.view", [
     "Consumption list loads",
     "Create consumption (internal)",
     "Create consumption from custody (Discipline + Milestone required)",
     "Cancel a consumption",
     "Cross-division consumption (permission-gated)",
     "Stock decrements after consume (view refreshes)",
   ]),
   ("Damaged Stock", "/warehouse/damaged-stock", "damaged_stock.*.view", [
     "On-hand + out-for-repair tabs load",
     "Send damaged stock for repair",
     "Return damaged from repair",
     "Disposition / write-off a damaged item",
   ]),
   ("Tools & Assets", "/warehouse/tools-assets", "tools.assets.view", [
     "Hub loads (Teams / Repair / Checks / History)",
     "Assign a serialized tool to a team",
     "Move a tool between same-division teams",
     "Return a tool",
     "Run a condition check (Good / Bad / Under-repair)",
     "Resolve repair -> Repaired or Scrap (posts unit cost to P&L)",
     "Cost hidden without tools.assets.cost.view",
     "Page is division-scoped (empty if user not in the division)",
   ]),
   ("Picture Transfer", "/warehouse/picture-transfer", "warehouse.transfer.simple", [
     "Send — create a pending picture transfer",
     "Receive — confirm an incoming picture transfer",
   ]),
   ("Warehouse Transfers (classic)", "/master-data/warehouses (Transfers)", "warehouse.transfers.view", [
     "Create a transfer",
     "Dispatch (issue) — Warehouse RP only",
     "Receive — Warehouse RP only",
     "Cancel / reject a transfer",
   ]),
   ("Stock Adjustments", "/master-data/warehouses (Adjustments)", "warehouse.adjustments.view", [
     "Request a stock adjustment",
     "Approve / reject an adjustment step",
     "Force-approve (Owner)",
     "On final approval, stock movement is booked",
     "Direct status write blocked (guarded)",
   ]),
   ("Inventory Checks", "/master-data/warehouses (Checks)", "warehouse.checks.view", [
     "Start an inventory check",
     "Save item counts",
     "Complete an assignment",
     "Approve / reject a check step",
     "Cancel a check",
   ]),
 ]),
 ("Cross-cutting", [
   ("Auth & session", "(all)", "-", [
     "Login with valid credentials",
     "Logout clears session",
     "Inactivity timeout logs out",
     "Session guard redirects unauthenticated users",
   ]),
   ("Navigation & permissions", "(all)", "route-permissions.ts", [
     "Nav dropdowns show only permitted items",
     "Direct URL to an unpermitted page is blocked (RoutePermissionGuard)",
     "System admin sees everything (bypass)",
     "A role with only one page lands there after login",
   ]),
   ("Division & financials", "(all)", "custom_access_token_hook", [
     "Division switcher changes scope of lists/reports",
     "Financials/claims work (JWT carries division claim — hook enabled)",
   ]),
   ("Responsive & stability", "(all)", "-", [
     "Mobile (<640) — hamburger / drawer nav works",
     "Tables collapse / horizontal-scroll on phone",
     "Dialogs full-screen on mobile, centered on desktop",
     "Selecting a dropdown value does not shift surrounding layout",
     "Dialog titles + column headers wrap (never truncated)",
   ]),
   ("Notifications", "(bell)", "-", [
     "Bell shows unread notifications",
     "Mark read / actioned / bulk-actioned",
   ]),
 ]),
]

# ---- palette ----
INK      = "16202E"
HEAD_BG  = "0E5960"
HEAD_FG  = "FFFFFF"
MOD_BG   = "DFF1F2"
MOD_FG   = "0A5960"
STAR_FG  = "A9640C"
PASS_BG  = "E6F6EC"; FAIL_BG = "FBEAE8"; BLOCK_BG = "FBF1DC"; NA_BG = "EDEFF2"
BORDER   = "D9E0E8"
ZEBRA    = "F6F9FB"

thin = Side(style="thin", color=BORDER)
grp  = Side(style="medium", color="0E5960")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(**k): return Font(name="Arial", **k)

wb = openpyxl.Workbook()

# ============ Sheet 1: Test Matrix ============
ws = wb.active
ws.title = "Test Matrix"
headers = ["Module", "Page", "Route", "Permission", "#", "Test Case", "Status", "Notes", "Tester", "Date"]
widths  = [16, 26, 30, 30, 5, 62, 12, 34, 12, 12]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = F(bold=True, color=HEAD_FG, size=11)
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26

r = 2
for mod, pages in MATRIX:
    for pidx, (page, route, perm, tests) in enumerate(pages):
        for i, test in enumerate(tests, 1):
            is_star = test.startswith("*")
            label = ("★ " + test[2:]) if is_star else test
            # thick top border on the first row of each module block (not the first row overall)
            top = grp if (pidx == 0 and i == 1 and r != 2) else thin
            vals = [mod, page, route, perm, i, label, "", "", "", ""]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=ci, value=v)
                c.border = Border(left=thin, right=thin, top=top, bottom=thin)
                if ci in (3, 4):
                    c.font = F(size=9, color="5C6A7B")
                else:
                    c.font = F(size=10, color=STAR_FG if (ci == 6 and is_star) else INK,
                               bold=(ci == 6 and is_star))
                c.alignment = Alignment(vertical="center", wrap_text=(ci in (6, 8)),
                                        horizontal="center" if ci == 5 else "left")
            r += 1

last = r - 1
TOTAL = last - 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{last}"

# data validation dropdown on Status (G)
dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True,
                    showDropDown=False)
dv.error = "Pick Pass, Fail, Blocked, or N/A"; dv.errorTitle = "Invalid status"
dv.prompt = "Pass / Fail / Blocked / N/A (leave blank = not tested)"; dv.promptTitle = "Status"
ws.add_data_validation(dv)
dv.add(f"G2:G{last}")

# conditional formatting — whole-row tint by status
rng = f"A2:J{last}"
ws.conditional_formatting.add(rng, FormulaRule(formula=['$G2="Pass"'],  fill=PatternFill("solid", fgColor=PASS_BG)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$G2="Fail"'],  fill=PatternFill("solid", fgColor=FAIL_BG)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$G2="Blocked"'],fill=PatternFill("solid", fgColor=BLOCK_BG)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$G2="N/A"'],   fill=PatternFill("solid", fgColor=NA_BG)))

# ============ Sheet 2: Summary ============
sm = wb.create_sheet("Summary", 0)
sm.sheet_view.showGridLines = False
for col, w in (("A", 26), ("B", 12), ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 14)):
    sm.column_dimensions[col].width = w

def put(cell, val, **font):
    c = sm[cell]; c.value = val; c.font = F(**font); return c

put("A1", "Shipping Build — QA Summary", bold=True, size=16, color=INK)
put("A2", "deploy/warehouse-shipping · back-office ERP", size=10, color="5C6A7B")

MX = "'Test Matrix'"
G = f"{MX}!$G$2:$G${last}"
B = f"{MX}!$B$2:$B${last}"

# overall block
put("A4", "Overall", bold=True, size=12, color=MOD_FG)
stats = [
    ("Total cases", f"=COUNTA({MX}!$F$2:$F${last})"),
    ("Pass",        f'=COUNTIF({G},"Pass")'),
    ("Fail",        f'=COUNTIF({G},"Fail")'),
    ("Blocked",     f'=COUNTIF({G},"Blocked")'),
    ("N/A",         f'=COUNTIF({G},"N/A")'),
    ("Not tested",  f"=B5-(B6+B7+B8+B9)"),
    ("% tested",    "=(B6+B7+B8+B9)/B5"),
    ("% pass (of tested)", "=IFERROR(B6/(B6+B7+B8),0)"),
]
row = 5
for lab, formula in stats:
    put(f"A{row}", lab, size=10, color=INK)
    c = sm[f"B{row}"]; c.value = formula; c.font = F(size=10, bold=True, color=INK)
    c.alignment = Alignment(horizontal="right")
    if lab.startswith("%"):
        c.number_format = "0.0%"
    row += 1
# color the status labels
sm["A6"].font = F(size=10, bold=True, color="15803D")
sm["A7"].font = F(size=10, bold=True, color="C0332A")
sm["A8"].font = F(size=10, bold=True, color="A9640C")

# per-module table
put("D4", "By module", bold=True, size=12, color=MOD_FG)
mheaders = ["Module", "Total", "Pass", "Fail", "Blocked", "N/A", "% tested"]
mcols = ["A", "B", "C", "D", "E", "F", "G"]  # laid out starting D? -> use D..J
start_col = 4  # D
mr = 5
for j, h in enumerate(mheaders):
    c = sm.cell(row=mr, column=start_col + j, value=h)
    c.font = F(bold=True, color=HEAD_FG, size=10)
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(horizontal="center")
    sm.column_dimensions[get_column_letter(start_col + j)].width = 12 if j else 26
mr += 1
for mod, pages in MATRIX:
    m = mod.replace('"', '""')
    sm.cell(row=mr, column=4, value=mod).font = F(size=10, color=INK)
    sm.cell(row=mr, column=5, value=f'=COUNTIF({B},"{m}")').font = F(size=10)
    sm.cell(row=mr, column=6, value=f'=COUNTIFS({B},"{m}",{G},"Pass")').font = F(size=10, color="15803D")
    sm.cell(row=mr, column=7, value=f'=COUNTIFS({B},"{m}",{G},"Fail")').font = F(size=10, color="C0332A")
    sm.cell(row=mr, column=8, value=f'=COUNTIFS({B},"{m}",{G},"Blocked")').font = F(size=10, color="A9640C")
    sm.cell(row=mr, column=9, value=f'=COUNTIFS({B},"{m}",{G},"N/A")').font = F(size=10)
    pct = sm.cell(row=mr, column=10,
                  value=f'=IFERROR((F{mr}+G{mr}+H{mr}+I{mr})/E{mr},0)'.replace("F", get_column_letter(6)))
    # build % tested = (pass+fail+blocked+na)/total using this row's C..F? recompute cleanly:
    pct.value = f"=IFERROR((F{mr}+G{mr}+H{mr}+I{mr})/E{mr},0)"
    pct.number_format = "0.0%"; pct.font = F(size=10, bold=True)
    mr += 1

# legend + example
lr = mr + 2
put(f"A{lr}", "How to use", bold=True, size=12, color=MOD_FG)
put(f"A{lr+1}", "Go to the 'Test Matrix' tab. For each row, pick a value in the Status column "
    "(dropdown: Pass / Fail / Blocked / N/A) and add a Notes / Tester / Date if useful. "
    "Rows tint by status and this Summary updates automatically. ★ rows test recent fixes.",
    size=10, color="5C6A7B")
sm[f"A{lr+1}"].alignment = Alignment(wrap_text=True, vertical="top")
sm.merge_cells(f"A{lr+1}:J{lr+1}")
sm.row_dimensions[lr+1].height = 42

put(f"A{lr+3}", "Example row (from Test Matrix):", bold=True, size=10, color=INK)
ex_h = ["Page", "Test Case", "Status", "Notes", "Tester", "Date"]
for j, h in enumerate(ex_h):
    c = sm.cell(row=lr+4, column=1 + j, value=h); c.font = F(bold=True, size=9, color="5C6A7B")
ex = ["Inventory", "Create item with a brand + origin variant", "Pass", "OK on staging", "Mr. Nijam", "2026-08-23"]
for j, v in enumerate(ex):
    c = sm.cell(row=lr+5, column=1 + j, value=v); c.font = F(size=9, color=INK)
    c.fill = PatternFill("solid", fgColor=PASS_BG)

import os
OUT = os.path.join(os.path.dirname(__file__), "shipping-qa-matrix.xlsx")
wb.save(OUT)
print("SAVED", OUT, "TOTAL_CASES", TOTAL, "LAST_ROW", last)
