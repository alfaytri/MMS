# Apply the 6 locked reorg changes + Tools-side rename to Inventory_Corrected.xlsx.
# FILE ONLY — does not touch the DB. Backs up first. Then verifies.
import openpyxl, shutil, datetime, sys, os, collections
sys.stdout.reconfigure(encoding="utf-8")
CORR=r"D:\ERP\Inventory Works\Inventory_Corrected.xlsx"
OUT=os.environ.get("REORG_OUT", CORR)   # save target; defaults to canonical file

def under(path, prefix):
    return path == prefix or path.startswith(prefix + " > ")

def route(path, typ):
    # Change 2: Home Automation -> Products (whole subtree, fixes the split)
    if under(path, "Home Automation"):
        return path, "products"
    # Change 6a: Safety & PPE -> Consumables
    if under(path, "Safety & PPE"):
        return path, "consumables"
    # Change 1: Pump Control -> Plumbing > Pressure Switch > Pump Control
    if under(path, "Pressure Switch > Pump Control"):
        return "Plumbing > " + path, "spare-parts"
    # Change 4: fold Pressure Washers into Cleaning Machines > Pressure Washer
    if path == "Pressure Washers":
        return "Cleaning Machines > Pressure Washer", "tools"
    if path.startswith("Pressure Washers > "):
        return "Cleaning Machines > Pressure Washer > " + path[len("Pressure Washers > "):], "tools"
    # Changes A + 5 + 6b: Cleaning Supplies subtree split
    if under(path, "Cleaning Supplies"):
        if path == "Cleaning Supplies":
            return "Cleaning Equipment", "tools"                       # A: rename Tools side
        if under(path, "Cleaning Supplies > Carpet & Stain Care"):
            return path, "consumables"                                 # 5: -> Consumables (rides Applicator/Bone Scraper = B, flagged)
        if under(path, "Cleaning Supplies > Chemicals"):
            return path, "consumables"                                 # 5
        if under(path, "Cleaning Supplies > Consumables > Garbage Bag"):
            return "Cleaning Supplies > " + path[len("Cleaning Supplies > Consumables > "):], "consumables"  # 6b
        if under(path, "Cleaning Supplies > Consumables > Table Roll"):
            return "Cleaning Supplies > " + path[len("Cleaning Supplies > Consumables > "):], "consumables"  # 6b
        # else stays in Tools (Air Freshener, Brooms, Pads, Wipers, + Bucket/Funnel/Spray Bottle)
        newp = "Cleaning Equipment" + path[len("Cleaning Supplies"):]
        # rename the leftover 'Consumables' wrapper -> 'Accessories' (owner pick)
        if newp == "Cleaning Equipment > Consumables" or newp.startswith("Cleaning Equipment > Consumables > "):
            newp = "Cleaning Equipment > Accessories" + newp[len("Cleaning Equipment > Consumables"):]
        return newp, "tools"
    return path, typ

ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
bak=fr"C:\Users\IT\AppData\Local\Temp\claude\D--MMS\cc79135c-8bed-48d0-94c3-4e8ab56529ff\scratchpad\source_backup\Inventory_Corrected.{ts}.bak.xlsx"
os.makedirs(os.path.dirname(bak),exist_ok=True)
shutil.copy(CORR,bak); print("backup:",bak,"\n")

wb=openpyxl.load_workbook(CORR)
ws=wb["Items"]; wc=wb["Categories"]
iix={c.value:i+1 for i,c in enumerate(ws[1])}
cixh={c.value:i+1 for i,c in enumerate(wc[1])}
ip,it=iix["Category Path"],iix["Type"]
cp,ct=cixh["Category Path"],cixh["Type"]

n_item=0
for r in range(2, ws.max_row+1):
    p=ws.cell(r,ip).value
    if p is None: continue
    p=str(p).strip(); t=str(ws.cell(r,it).value or "").strip()
    np,nt=route(p,t)
    if (np,nt)!=(p,t):
        ws.cell(r,ip).value=np; ws.cell(r,it).value=nt; n_item+=1

n_cat=0
for r in range(2, wc.max_row+1):
    p=wc.cell(r,cp).value
    if p is None: continue
    p=str(p).strip(); t=str(wc.cell(r,ct).value or "").strip()
    np,nt=route(p,t)
    if (np,nt)!=(p,t):
        wc.cell(r,cp).value=np; wc.cell(r,ct).value=nt; n_cat+=1

now_paths={str(wc.cell(r,cp).value).strip() for r in range(2,wc.max_row+1) if wc.cell(r,cp).value is not None}
to_add=[]
if "Plumbing > Pressure Switch" not in now_paths: to_add.append(("Plumbing > Pressure Switch","spare-parts"))
if "Cleaning Supplies" not in now_paths:          to_add.append(("Cleaning Supplies","consumables"))
for path,typ in to_add:
    row=[None]*wc.max_column; row[cp-1]=path; row[ct-1]=typ
    wc.append(row); print("added category:",typ,"|",path)

wb.save(OUT)
print(f"\nsaved -> {OUT}\nrewrote {n_item} item rows + {n_cat} category rows; added {len(to_add)} categories")

# ---------- VERIFY (re-open saved file) ----------
print("\n================ AFTER (affected branches) ================")
wb2=openpyxl.load_workbook(OUT); ws2=wb2["Items"]; wc2=wb2["Categories"]
i2={c.value:i for i,c in enumerate(ws2[1])}; c2={c.value:i for i,c in enumerate(wc2[1])}
cats=[]
for r in range(2,wc2.max_row+1):
    p=wc2.cell(r,c2["Category Path"]+1).value
    if p: cats.append((str(p).strip(), str(wc2.cell(r,c2["Type"]+1).value or "").strip()))
items=collections.defaultdict(int)
tabcount=collections.Counter()
for r in range(2,ws2.max_row+1):
    p=ws2.cell(r,i2["Category Path"]+1).value
    if not p: continue
    items[str(p).strip()]+=1
    tabcount[str(ws2.cell(r,i2["Type"]+1).value or "").strip()]+=1

def show(title, prefixes):
    print(f"\n----- {title} -----")
    for p,t in sorted(cats):
        if any(p==pr or p.startswith(pr+" > ") for pr in prefixes):
            d=p.count(" > "); print(f"{'  '*d}{p.split(' > ')[-1]}  <{t}>"+(f"  ({items.get(p,0)})" if items.get(p,0) else ""))

show("Home Automation (now Products)", ["Home Automation"])
show("Plumbing > Pressure Switch (Pump Control moved in)", ["Plumbing > Pressure Switch"])
show("top-level Pressure Switch (HP/LP remain)", ["Pressure Switch"])
show("Cleaning Machines (Pressure Washers folded in)", ["Cleaning Machines"])
show("Cleaning Equipment (renamed Tools side)", ["Cleaning Equipment"])
show("Cleaning Supplies (new Consumables parent)", ["Cleaning Supplies"])
show("Safety & PPE (now Consumables)", ["Safety & PPE"])

print("\nitems per tab:", dict(tabcount))
# parent/child type consistency (report mismatches)
tmap={p:t for p,t in cats}
mism=[]
for p,t in cats:
    if " > " in p:
        par=p.rsplit(" > ",1)[0]
        if par in tmap and tmap[par]!=t: mism.append((p,t,tmap[par]))
print(f"\nparent/child type mismatches remaining: {len(mism)}")
for p,t,pt in mism: print(f"   {p}  <{t}>  (parent is <{pt}>)")
