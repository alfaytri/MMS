#!/usr/bin/env python3
"""Reload the corrected inventory catalog into new-prod.
Default = DRY RUN (parses + reports, writes nothing). Pass --commit to write
(atomic: wipe + reload in one transaction; rollback on any error).
Source: D:\\ERP\\Inventory Works\\Inventory_Corrected.xlsx (Items + Categories).
"""
import os, sys, json, datetime
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

COMMIT = '--commit' in sys.argv
XLSX = r"D:\ERP\Inventory Works\Inventory_Corrected.xlsx"
DB = os.environ['NEW_DB_URL']

DIV = {  # division name -> uuid (new-prod)
  'Kitchen':'8215e9b4-0292-4914-a584-c2b26610378f',
  'Maintenance':'b0dfc0f4-fe96-4582-a517-276aa901d541',
  'MEP':'2e0018bf-fcb3-481f-9aae-820cc77c060e',
  'Pest Control & Cleaning':'c6070d16-a7e4-4ee8-a8ad-0dae5ce3d583',
  'Trading':'1b821883-c99e-4177-8144-1c5c1204b007',
}
DEFAULT_CONTAINER = {  # division name -> sub_container uuid (new-prod)
  'Trading':'1a09f294-eb9f-4324-b3d6-d4dc26f2de38',
  'Maintenance':'de100a9d-be78-42c8-b6f2-aff11b258576',
  'MEP':'3fbd13e6-96b2-483f-84f1-6daf517dc98f',
  'Pest Control & Cleaning':'25a40a79-e5fa-4a3c-a68e-ed95d2d69ea1',
  'Kitchen':'81610789-4530-4bc2-ab49-f71b73ff3fba',
}
ORIGIN_ALIAS = {'usa':'united states','uk':'united kingdom','uae':'united arab emirates',
                'scotland':'united kingdom'}  # Scotland -> UK (flagged below)

def split_list(v):
    if v is None: return []
    s = str(v).strip()
    if not s or s.lower() == 'none': return []
    return [p.strip() for p in s.split(',') if p.strip()]

wb = load_workbook(XLSX, read_only=True, data_only=True)
irows = list(wb['Items'].iter_rows(values_only=True)); ih = {h:i for i,h in enumerate(irows[0])}
crows = list(wb['Categories'].iter_rows(values_only=True)); ch = {h:i for i,h in enumerate(crows[0])}
items = irows[1:]; cats = crows[1:]

conn = psycopg2.connect(DB); cur = conn.cursor()
cur.execute("select lower(name), id from country_codes"); country_by_name = dict(cur.fetchall())
def origin_id(o):
    k = o.strip().lower()
    return country_by_name.get(k) or country_by_name.get(ORIGIN_ALIAS.get(k,''))
cur.execute("select lower(trim(name)), id from brands"); brand_by_name = dict(cur.fetchall())

warn = []

# ---- categories: (type, path) -> planned row; insert parent-first (by depth) ----
cat_plan = {}  # (type, path_lower) -> {name,type,parent_path_lower,div}
for r in cats:
    path = str(r[ch['Category Path']]).strip(); typ = r[ch['Type']]; div = r[ch['Default Container (Division)']]
    segs = [s.strip() for s in path.split('>') if s.strip()]
    plow = ' > '.join(s.lower() for s in segs)
    parent_plow = ' > '.join(s.lower() for s in segs[:-1]) if len(segs) > 1 else None
    divname = None if (div is None or str(div).strip().lower()=='none') else str(div).strip()
    if divname and divname not in DEFAULT_CONTAINER: warn.append(f"cat default-container div not mapped: {divname!r} ({path})")
    cat_plan[(typ, plow)] = {'name':segs[-1], 'type':typ, 'parent_plow':parent_plow,
                             'div':divname, 'depth':len(segs), 'path':path}

# Each category path is unique to one type (verified: 0 paths under >1 type), so a
# path maps to exactly one category. Parents/items resolve by PATH (the parent's
# own type may differ from the child's — mixed-type split branches).
by_path = {plow: (typ, plow) for (typ, plow) in cat_plan}

# ---- items + variants ----
item_plan = []   # {name,sku,unit,spec,cat_key,div_ids}
variant_plan = []  # index into item_plan -> list of {brand,origin}
for r in items:
    name = str(r[ih['Item Name']]).strip(); typ = r[ih['Type']]; unit = r[ih['Unit']]
    path = str(r[ih['Category Path']]).strip()
    spec = r[ih['Specifications']]
    brands = split_list(r[ih['Brand(s)']]); origins = split_list(r[ih['Origin(s)']])
    branches = split_list(r[ih['Branches']])
    segs = [s.strip() for s in path.split('>') if s.strip()]
    item_plow = ' > '.join(s.lower() for s in segs)
    ckey = by_path.get(item_plow)
    if ckey is None: warn.append(f"item category path not found: {name!r} -> {path}")
    elif ckey[0] != typ: warn.append(f"item Type {typ!r} != category type {ckey[0]!r}: {name!r} ({path})")
    div_ids = []
    for b in branches:
        if b in DIV: div_ids.append(DIV[b])
        else: warn.append(f"item branch div not mapped: {b!r} ({name})")
    # brand/origin pairing
    vlist = []
    if not brands:
        pass  # item with no brand -> no variant (rare; flag)
    elif origins and len(origins) == len(brands):
        vlist = [{'brand':b,'origin':o} for b,o in zip(brands,origins)]      # paired
    elif len(brands) == 1 and origins:
        vlist = [{'brand':brands[0],'origin':o} for o in origins]            # 1 brand from N origins -> N variants
    elif len(origins) == 1:
        vlist = [{'brand':b,'origin':origins[0]} for b in brands]            # N brands, all from 1 origin
    elif not origins:
        vlist = [{'brand':b,'origin':None} for b in brands]                  # N brands, no origin
    else:
        warn.append(f"brand/origin count mismatch (both >1, unequal): {name!r} brands={brands} origins={origins}")
        vlist = [{'brand':b,'origin':(origins[i] if i < len(origins) else None)} for i,b in enumerate(brands)]
    if not vlist: warn.append(f"item has NO brand -> 0 variants: {name!r} ({path})")
    item_plan.append({'name':name,'unit':unit,'spec':(str(spec).strip() if spec else None),
                      'cat_key':ckey,'div_ids':div_ids})
    variant_plan.append(vlist)

# origin/brand stats
all_origins = set(); unmatched_origins = set(); total_variants = 0; brands_needed = set()
for vl in variant_plan:
    for v in vl:
        total_variants += 1
        if v['origin']:
            all_origins.add(v['origin'])
            if origin_id(v['origin']) is None: unmatched_origins.add(v['origin'])
        if v['brand']: brands_needed.add(v['brand'].strip().lower())
new_brands = [b for b in brands_needed if b not in brand_by_name]

print("=== DRY RUN ===" if not COMMIT else "=== COMMIT MODE ===")
print(f"categories in sheet: {len(cat_plan)}")
print(f"items: {len(item_plan)}  |  planned variants: {total_variants}")
print(f"brands: {len(brands_needed)} needed, {len(new_brands)} new to create")
print(f"origins used: {sorted(all_origins)}")
print(f"UNMATCHED origins (would be NULL unless aliased): {sorted(unmatched_origins) or 'none'}")
print(f"Scotland -> United Kingdom alias active: {'scotland' in ORIGIN_ALIAS}")
from collections import Counter
divc = Counter()
for ip in item_plan:
    divc[len(ip['div_ids'])] += 1
print(f"item division-scoping (n_divisions -> item count): {dict(divc)}")
print(f"items with 0 divisions (would be 'All'-only): {divc[0]}")
print(f"WARNINGS ({len(warn)}):")
for w in warn[:25]: print("  -", w)
if len(warn) > 25: print(f"  ... +{len(warn)-25} more")

if not COMMIT:
    print("\n(dry run — nothing written. Re-run with --commit to load.)")
    sys.exit(0)

# ================= COMMIT: atomic wipe + reload =================
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
try:
    cur.execute("select count(*) from inventory_stock_movements"); mv = cur.fetchone()[0]
    if mv != 0:
        raise SystemExit(f"ABORT: new-prod has {mv} stock movements — not a safe wipe target.")
    # snapshot current categories (items/variants are 0)
    cur.execute("select id::text, name_en, type, parent_id::text, default_sub_container_id::text from inventory_categories")
    snap = cur.fetchall()
    with open(rf"C:\Users\IT\AppData\Local\Temp\claude\D--MMS\6deb37d1-86b3-4edd-85aa-944d192bcf69\scratchpad\prod_categories_snapshot_{ts}.json","w",encoding="utf-8") as f:
        json.dump(snap, f)
    print(f"snapshot: {len(snap)} categories saved")

    # wipe (0 variants/items; 447 categories)
    cur.execute("delete from inventory_item_brand_variants")
    cur.execute("delete from inventory_items")
    cur.execute("delete from inventory_categories")

    # create brands (batch)
    if new_brands:
        for row in execute_values(cur, "insert into brands(name) values %s returning id, lower(trim(name))",
                                  [(b,) for b in new_brands], page_size=1000, fetch=True):
            brand_by_name[row[1]] = row[0]

    # categories — insert per depth level (parents first), batched; map ids back by
    # (parent_id, lower(name), type) natural key (not RETURNING order).
    cat_id = {}
    depths = sorted(set(cat_plan[k]['depth'] for k in cat_plan))
    for depth in depths:
        keys = [k for k in cat_plan if cat_plan[k]['depth'] == depth]
        vals = []
        for k in keys:
            c = cat_plan[k]
            pkey = by_path.get(c['parent_plow']) if c['parent_plow'] else None
            parent_id = cat_id.get(pkey)
            dsc = DEFAULT_CONTAINER.get(c['div']) if c['div'] else None
            vals.append((c['name'], c['type'], parent_id, dsc, 'active', 0))
        execute_values(cur,
            "insert into inventory_categories(name_en,type,parent_id,default_sub_container_id,status,sort_order) values %s",
            vals, template="(%s,%s,%s,%s,%s,%s)", page_size=10000)
        cur.execute("select id::text, coalesce(parent_id::text,''), lower(trim(name_en)), type from inventory_categories")
        lookup = {(p, n, t): i for i, p, n, t in cur.fetchall()}
        for k in keys:
            c = cat_plan[k]
            pkey = by_path.get(c['parent_plow']) if c['parent_plow'] else None
            pid = cat_id.get(pkey)
            cat_id[k] = lookup[(str(pid) if pid else '', c['name'].strip().lower(), c['type'])]

    # items — batched; map ids by (category_id, lower(name))
    ivals = [(cat_id.get(ip['cat_key']), ip['name'], ip['name'], ip['unit'], ip['spec'],
              ip['div_ids'] or None, 'active', 0) for ip in item_plan]
    execute_values(cur,
        "insert into inventory_items(category_id,name_en,sku,unit,specification,shared_with_division_ids,status,sort_order) values %s",
        ivals, template="(%s,%s,%s,%s,%s,%s::uuid[],%s,%s)", page_size=10000)
    cur.execute("select id::text, category_id::text, lower(trim(name_en)) from inventory_items")
    imap = {(c, n): i for i, c, n in cur.fetchall()}
    item_ids = [imap[(str(cat_id.get(ip['cat_key'])), ip['name'].strip().lower())] for ip in item_plan]

    # variants — batched
    vvals = []
    for i, vl in enumerate(variant_plan):
        for v in vl:
            vvals.append((item_ids[i], v['brand'], brand_by_name.get(v['brand'].strip().lower()),
                          origin_id(v['origin']) if v['origin'] else None, 0, 'active', 0))
    execute_values(cur,
        "insert into inventory_item_brand_variants(item_id,brand,brand_id,country_id,stock_level,status,sort_order) values %s",
        vvals, template="(%s,%s,%s,%s,%s,%s,%s)", page_size=10000)
    vcount = len(vvals)

    conn.commit()
    print(f"COMMITTED: {len(cat_id)} categories, {len(item_ids)} items, {vcount} variants.")
except Exception as e:
    conn.rollback()
    print("ROLLED BACK:", e)
    raise
