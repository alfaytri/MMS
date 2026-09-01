#!/usr/bin/env python3
"""new-prod inventory: wipe (catalog + opening-stock artifacts) and reload the
UPDATED catalog at ZERO quantity, PRESERVING current variant prices matched by
(type, category-path, item-name, brand).

Modes:
  (default)   DRY RUN  — parse + report + price-match preview + uniqueness checks; writes nothing.
  --rehearse  do the full wipe+reload+verify in one tx, then ROLL BACK (proves the path).
  --commit    do it for real and COMMIT.

A full pg_dump backup must be taken separately BEFORE --commit.
Source file: env INVENTORY_XLSX (default = the updated file in Downloads).
DB: env NEW_DB_URL (new-prod).
"""
import os, sys, json, datetime
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

MODE = 'commit' if '--commit' in sys.argv else 'rehearse' if '--rehearse' in sys.argv else 'dry'
XLSX = os.environ.get('INVENTORY_XLSX', r"C:\Users\IT\Downloads\Inventory_Updated (1).xlsx")
DB = os.environ['NEW_DB_URL']
# snapshot dir: env override, else alongside this script
SNAP_DIR = os.environ.get('SNAP_DIR', os.path.dirname(os.path.abspath(__file__)))

DIV = {  # division name -> uuid (new-prod)
  'Kitchen':'8215e9b4-0292-4914-a584-c2b26610378f',
  'Maintenance':'b0dfc0f4-fe96-4582-a517-276aa901d541',
  'MEP':'2e0018bf-fcb3-481f-9aae-820cc77c060e',
  'Pest Control & Cleaning':'c6070d16-a7e4-4ee8-a8ad-0dae5ce3d583',
  'Trading':'1b821883-c99e-4177-8144-1c5c1204b007',
}
DEFAULT_CONTAINER = {  # division name -> sub_container uuid (new-prod)
  'Trading':'1a09f294-eb9f-4324-b3d6-d4dc26f2de38',
  'Maintenance':'de100a9d-be78-42c8-b6f2-aff11b258576',
  'MEP':'3fbd13e6-96b2-483f-84f1-6daf517dc98f',
  'Pest Control & Cleaning':'25a40a79-e5fa-4a3c-a68e-ed95d2d69ea1',
  'Kitchen':'81610789-4530-4bc2-ab49-f71b73ff3fba',
}
ORIGIN_ALIAS = {'usa':'united states','uk':'united kingdom','uae':'united arab emirates',
                'scotland':'united kingdom'}

def split_list(v):
    if v is None: return []
    s = str(v).strip()
    if not s or s.lower() == 'none': return []
    return [p.strip() for p in s.split(',') if p.strip()]

# ---------- parse workbook ----------
wb = load_workbook(XLSX, read_only=True, data_only=True)
irows = list(wb['Items'].iter_rows(values_only=True)); ih = {h:i for i,h in enumerate(irows[0])}
crows = list(wb['Categories'].iter_rows(values_only=True)); ch = {h:i for i,h in enumerate(crows[0])}
items = irows[1:]; cats = crows[1:]

conn = psycopg2.connect(DB); cur = conn.cursor()
cur.execute("select lower(name), id from country_codes"); country_by_name = dict(cur.fetchall())
def origin_id(o):
    if not o: return None
    k = o.strip().lower()
    return country_by_name.get(k) or country_by_name.get(ORIGIN_ALIAS.get(k,''))
cur.execute("select lower(trim(name)), id from brands"); brand_by_name = dict(cur.fetchall())

warn = []

# categories: (type, path_lower) -> plan.
# Every ancestor prefix is materialized (the file lists some leaf paths whose
# intermediate parents are not their own rows). This matches the app importer,
# which auto-creates missing ancestor categories from each path. Synthesized
# ancestors get name/type from the path and no default division; a prefix that
# IS explicitly listed keeps its own division.
cat_plan = {}
def _add_cat(typ, segs, divname):
    for d in range(1, len(segs) + 1):
        pre = segs[:d]
        plow = ' > '.join(s.lower() for s in pre)
        parent_plow = ' > '.join(s.lower() for s in pre[:-1]) if d > 1 else None
        key = (typ, plow); is_leaf = (d == len(segs))
        if key not in cat_plan:
            cat_plan[key] = {'name':pre[-1], 'type':typ, 'parent_plow':parent_plow,
                             'div':(divname if is_leaf else None), 'depth':d,
                             'path':' > '.join(pre), 'synth':(not is_leaf)}
        elif is_leaf:
            cat_plan[key]['synth'] = False
            if divname and not cat_plan[key]['div']:
                cat_plan[key]['div'] = divname
for r in cats:
    path = str(r[ch['Category Path']]).strip(); typ = r[ch['Type']]; div = r[ch['Default Container (Division)']]
    segs = [s.strip() for s in path.split('>') if s.strip()]
    divname = None if (div is None or str(div).strip().lower()=='none') else str(div).strip()
    if divname and divname not in DEFAULT_CONTAINER: warn.append(f"cat default-container div not mapped: {divname!r} ({path})")
    _add_cat(typ, segs, divname)
_synth = sorted(c['path'] for c in cat_plan.values() if c.get('synth'))
by_path = {plow: (typ, plow) for (typ, plow) in cat_plan}

# items + variants
item_plan = []   # {name,unit,spec,cat_key,div_ids,typ,plow}
variant_plan = []  # per item -> [{brand,origin}]
for r in items:
    name = str(r[ih['Item Name']]).strip(); typ = r[ih['Type']]; unit = r[ih['Unit']]
    path = str(r[ih['Category Path']]).strip()
    spec = r[ih['Specifications']]
    brands = split_list(r[ih['Brand(s)']]); origins = split_list(r[ih['Origin(s)']])
    branches = split_list(r[ih['Branches']])
    segs = [s.strip() for s in path.split('>') if s.strip()]
    item_plow = ' > '.join(s.lower() for s in segs)
    ckey = by_path.get(item_plow)
    if ckey is None: warn.append(f"item category path not found: {name!r} -> {path}")
    elif ckey[0] != typ: warn.append(f"item Type {typ!r} != category type {ckey[0]!r}: {name!r} ({path})")
    div_ids = []
    for b in branches:
        if b in DIV: div_ids.append(DIV[b])
        else: warn.append(f"item branch div not mapped: {b!r} ({name})")
    if not brands:
        vlist = []
    elif origins and len(origins) == len(brands):
        vlist = [{'brand':b,'origin':o} for b,o in zip(brands,origins)]
    elif len(brands) == 1 and origins:
        vlist = [{'brand':brands[0],'origin':o} for o in origins]
    elif len(origins) == 1:
        vlist = [{'brand':b,'origin':origins[0]} for b in brands]
    elif not origins:
        vlist = [{'brand':b,'origin':None} for b in brands]
    else:
        warn.append(f"brand/origin count mismatch (both >1, unequal): {name!r} brands={brands} origins={origins}")
        vlist = [{'brand':b,'origin':(origins[i] if i < len(origins) else None)} for i,b in enumerate(brands)]
    if not vlist: warn.append(f"item has NO brand -> 0 variants: {name!r} ({path})")
    item_plan.append({'name':name,'unit':unit,'spec':(str(spec).strip() if spec else None),
                      'cat_key':ckey,'div_ids':div_ids,'typ':(ckey[0] if ckey else typ),'plow':item_plow})
    variant_plan.append(vlist)

# ---------- price preservation: snapshot BEFORE wipe ----------
cur.execute("select id::text, name_en, type, coalesce(parent_id::text,'') from inventory_categories")
oc = {r[0]: (r[1], r[2], r[3]) for r in cur.fetchall()}
def old_path(cid):
    segs, seen = [], set()
    while cid and cid in oc and cid not in seen:
        seen.add(cid); name, _t, parent = oc[cid]
        segs.append((name or '').strip().lower()); cid = parent or None
    segs.reverse(); return ' > '.join(segs)
def old_type(cid): return oc.get(cid, (None,None,None))[1]

def pkey(typ, plow, name, brand):
    return f"{typ}::{plow}|{(name or '').strip().lower()}|{(brand or '').strip().lower()}"
def ikey(typ, plow, name):
    return f"{typ}::{plow}|{(name or '').strip().lower()}"

cur.execute("""select v.brand, i.category_id::text, i.name_en, v.cost_price, v.selling_price, v.average_cost
               from inventory_item_brand_variants v join inventory_items i on i.id=v.item_id""")
old_price = {}
for brand, cat_id, name, cost, sell, avg in cur.fetchall():
    k = pkey(old_type(cat_id), old_path(cat_id), name, brand)
    old_price.setdefault(k, {'cost':cost,'sell':sell,'avg':avg})
cur.execute("select category_id::text, name_en, cost_price from inventory_items")
old_item_cost = {}
for cat_id, name, cost in cur.fetchall():
    old_item_cost.setdefault(ikey(old_type(cat_id), old_path(cat_id), name), cost)

# ---------- match preview + uniqueness checks ----------
from collections import defaultdict
_nb_prices = defaultdict(set)   # name|brand -> set of distinct (cost,sell)
_nb_val = {}
for k,v in old_price.items():
    _tp, rest = k.split('::',1); _path, nb = rest.split('|',1)  # nb = name|brand
    _nb_prices[nb].add((None if v['cost'] is None else float(v['cost']),
                        None if v['sell'] is None else float(v['sell'])))
    _nb_val.setdefault(nb, v)
old_nb_unique = {nb: _nb_val[nb] for nb, s in _nb_prices.items() if len(s) == 1}

tot_var = 0; matched = 0; fb_ok = 0; fb_ambig = 0
for i, vl in enumerate(variant_plan):
    ip = item_plan[i]
    for v in vl:
        tot_var += 1
        if pkey(ip['typ'], ip['plow'], ip['name'], v['brand']) in old_price:
            matched += 1
        else:
            nb = f"{ip['name'].strip().lower()}|{(v['brand'] or '').strip().lower()}"
            if nb in old_nb_unique: fb_ok += 1
            elif nb in _nb_prices: fb_ambig += 1

# item uniqueness (uq_inventory_items_name_category = category_id + lower(name))
item_dupes = {}
for ip in item_plan:
    k = (ip['cat_key'], ip['name'].strip().lower())
    item_dupes[k] = item_dupes.get(k,0)+1
item_dupes = {k:c for k,c in item_dupes.items() if c>1 and k[0] is not None}
# variant uniqueness (uq_iibv_item_brand_origin = item_id + brand_id + country_id, NULLS NOT DISTINCT)
var_dupes = []
for i, vl in enumerate(variant_plan):
    seen = {}
    for v in vl:
        bid = brand_by_name.get((v['brand'] or '').strip().lower())
        cid = origin_id(v['origin'])
        key = (bid, cid)
        seen[key] = seen.get(key,0)+1
    for key,c in seen.items():
        if c>1: var_dupes.append((item_plan[i]['name'], item_plan[i]['plow'], key, c))

print(f"=== MODE: {MODE.upper()} ===")
print(f"source: {XLSX}")
print(f"planned: {len(cat_plan)} categories ({len(_synth)} synthesized ancestors) | {len(item_plan)} items | {tot_var} variants")
if _synth: print(f"  synthesized ancestor categories: {_synth}")
_brands_needed = {(v['brand'] or '').strip().lower() for vl in variant_plan for v in vl if v['brand']}
print(f"brands used: {len(_brands_needed)} distinct | new to create: {len([b for b in _brands_needed if b not in brand_by_name])}")
print(f"PRICE MATCH (keep-prices): {matched}/{tot_var} strict (type+path+item+brand); "
      f"+{fb_ok} via unambiguous name+brand fallback; {fb_ambig} ambiguous-skip; "
      f"{tot_var-matched-fb_ok-fb_ambig} no old price => {matched+fb_ok}/{tot_var} will keep a price.")
print(f"item uniqueness violations (category+name): {len(item_dupes)}  {list(item_dupes.items())[:5]}")
print(f"variant uniqueness violations (item+brand+origin): {len(var_dupes)}  {var_dupes[:5]}")
print(f"WARNINGS ({len(warn)}):")
for w in warn[:20]: print("  -", w)
if len(warn) > 20: print(f"  ... +{len(warn)-20} more")

if item_dupes or var_dupes:
    print("\n!! ABORT-WORTHY: unique-index collisions would fail the batch insert. Fix source before commit.")
    if MODE != 'dry':
        conn.rollback(); raise SystemExit("Refusing to write with unique-index collisions present.")

if MODE == 'dry':
    print("\n(dry run — nothing written. --rehearse to prove end-to-end + rollback; --commit to load.)")
    sys.exit(0)

# ================= WRITE (rehearse=rollback / commit=commit) =================
ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
try:
    # safety guard: no real transactions
    REAL = ['cogs_entries','sale_order_lines','po_line_items','consumption_entries','consumption_lines',
            'receival_items','bill_line_items','invoice_line_items','return_lines','inventory_damaged_stock',
            'stock_adjustments','inventory_check_items']
    for t in REAL:
        cur.execute(f"select count(*) from {t}"); n=cur.fetchone()[0]
        if n: raise SystemExit(f"ABORT: {t} has {n} rows — real transactions present, unsafe to wipe.")

    # snapshot old prices (audit / recovery)
    with open(rf"{SNAP_DIR}\newprod_old_prices_{ts}.json","w",encoding="utf-8") as f:
        json.dump({k:{'cost':(float(v['cost']) if v['cost'] is not None else None),
                      'sell':(float(v['sell']) if v['sell'] is not None else None),
                      'avg':(float(v['avg']) if v['avg'] is not None else None)}
                   for k,v in old_price.items()}, f)
    print(f"old-price snapshot: {len(old_price)} keys saved")

    # ---- WIPE (replica: FK + referential actions + triggers off) ----
    cur.execute("set local session_replication_role = replica")
    for t in ['warehouse_stock_allocations','warehouse_reorder_points','warehouse_transfer_items',
              'warehouse_transfers','inventory_stock_movements','fifo_cost_layers','warehouse_stock_summary',
              'inventory_item_brand_variants','inventory_item_attributes','inventory_item_divisions',
              'inventory_items','inventory_categories']:
        cur.execute(f"delete from {t}")
    cur.execute("set local session_replication_role = default")  # triggers+FK back ON for inserts

    # ---- brands (create any missing) ----
    new_brands = [b for b in {(v['brand'] or '').strip().lower() for vl in variant_plan for v in vl}
                  if b and b not in brand_by_name]
    if new_brands:
        for row in execute_values(cur, "insert into brands(name) values %s returning id, lower(trim(name))",
                                  [(b,) for b in new_brands], page_size=1000, fetch=True):
            brand_by_name[row[1]] = row[0]

    # ---- categories: parent-first by depth; map ids by natural key ----
    cat_id = {}
    for depth in sorted(set(c['depth'] for c in cat_plan.values())):
        keys = [k for k in cat_plan if cat_plan[k]['depth'] == depth]
        vals = []
        for k in keys:
            c = cat_plan[k]
            pkey_ = by_path.get(c['parent_plow']) if c['parent_plow'] else None
            parent_id = cat_id.get(pkey_)
            dsc = DEFAULT_CONTAINER.get(c['div']) if c['div'] else None
            vals.append((c['name'], c['type'], parent_id, dsc, 'active', 0))
        execute_values(cur,
            "insert into inventory_categories(name_en,type,parent_id,default_sub_container_id,status,sort_order) values %s",
            vals, template="(%s,%s,%s,%s,%s,%s)", page_size=10000)
        cur.execute("select id::text, coalesce(parent_id::text,''), lower(trim(name_en)), type from inventory_categories")
        lookup = {(p, n, t): i for i, p, n, t in cur.fetchall()}
        for k in keys:
            c = cat_plan[k]
            pkey_ = by_path.get(c['parent_plow']) if c['parent_plow'] else None
            pid = cat_id.get(pkey_)
            cat_id[k] = lookup[(str(pid) if pid else '', c['name'].strip().lower(), c['type'])]

    # ---- items (with preserved item cost) ----
    ivals = []
    for ip in item_plan:
        icost = old_item_cost.get(ikey(ip['typ'], ip['plow'], ip['name']))
        ivals.append((cat_id.get(ip['cat_key']), ip['name'], ip['name'], ip['unit'], ip['spec'],
                      icost if icost is not None else 0, 'active', 0))
    execute_values(cur,
        "insert into inventory_items(category_id,name_en,sku,unit,specification,cost_price,status,sort_order) values %s",
        ivals, template="(%s,%s,%s,%s,%s,%s,%s,%s)", page_size=10000)
    cur.execute("select id::text, category_id::text, lower(trim(name_en)) from inventory_items")
    imap = {(c, n): i for i, c, n in cur.fetchall()}
    item_ids = [imap[(str(cat_id.get(ip['cat_key'])), ip['name'].strip().lower())] for ip in item_plan]

    # ---- item -> divisions ----
    aivals = []
    for i, ip in enumerate(item_plan):
        for did in (ip['div_ids'] or []):
            aivals.append((item_ids[i], did, cat_id.get(ip['cat_key'])))
    if aivals:
        execute_values(cur,
            "insert into inventory_item_divisions(item_id,division_id,category_id) values %s "
            "on conflict (item_id,division_id) do nothing",
            aivals, template="(%s::uuid,%s::uuid,%s::uuid)", page_size=10000)

    # ---- variants (ZERO qty, PRESERVED prices) ----
    vvals = []
    for i, vl in enumerate(variant_plan):
        ip = item_plan[i]
        for v in vl:
            pr = old_price.get(pkey(ip['typ'], ip['plow'], ip['name'], v['brand']))
            if not pr:  # path changed → fall back to unambiguous name+brand price
                pr = old_nb_unique.get(f"{ip['name'].strip().lower()}|{(v['brand'] or '').strip().lower()}")
            cost = pr['cost'] if pr else 0
            sell = pr['sell'] if pr else 0
            avg  = pr['avg']  if pr else None
            vvals.append((item_ids[i], v['brand'], brand_by_name.get((v['brand'] or '').strip().lower()),
                          origin_id(v['origin']), 0, cost, sell, avg, 'active', 0))
    execute_values(cur,
        "insert into inventory_item_brand_variants"
        "(item_id,brand,brand_id,country_id,stock_level,cost_price,selling_price,average_cost,status,sort_order) values %s",
        vvals, template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", page_size=10000)

    # ---- verify ----
    def one(q):
        cur.execute(q); return cur.fetchone()[0]
    v_cats = one("select count(*) from inventory_categories")
    v_items = one("select count(*) from inventory_items")
    v_vars = one("select count(*) from inventory_item_brand_variants")
    v_divs = one("select count(*) from inventory_item_divisions")
    v_fifo = one("select count(*) from fifo_cost_layers")
    v_mov  = one("select count(*) from inventory_stock_movements")
    v_summ = one("select count(*) from warehouse_stock_summary")
    v_qty  = one("select coalesce(sum(stock_level),0) from inventory_item_brand_variants")
    v_priced = one("select count(*) from inventory_item_brand_variants where coalesce(selling_price,0)<>0 or coalesce(cost_price,0)<>0")
    print("\n--- VERIFY (in-tx) ---")
    print(f"categories={v_cats} items={v_items} variants={v_vars} item_divisions={v_divs}")
    print(f"ZERO-QTY check: fifo_layers={v_fifo} stock_movements={v_mov} stock_summary={v_summ} sum(stock_level)={v_qty}")
    print(f"variants with a non-zero price: {v_priced}/{v_vars}")

    if MODE == 'rehearse':
        conn.rollback()
        print("\nREHEARSE COMPLETE — rolled back (no changes persisted).")
    else:
        conn.commit()
        print("\nCOMMITTED to new-prod.")
except Exception as e:
    conn.rollback()
    print("ROLLED BACK:", repr(e))
    raise
finally:
    cur.close(); conn.close()
